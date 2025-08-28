# FILENAME:ImportEvents.py
# AUTHOR:Jonathan Shambaugh
# PURPOSE: To extract the comments given in a Toyopuc project
#   and write them to the corresponding address in the template for easy event importing.
# NOTES: See the github repository for more info.
#   https:\\github.com/jkernal/EDC_Events_tool
# VERSION: v2.0.0
# START DATE: 17 Oct 22

from mmap import ACCESS_READ, mmap
from os import R_OK, W_OK, X_OK, access, getcwd, listdir, system
from pathlib import Path
from re import fullmatch, search
from shutil import copy
from subprocess import check_call, check_output, run
from sys import executable, version
from time import perf_counter
from datetime import datetime

#import utils
from utils.load_config import load_config
from utils.logger_setup import setup_logging

WRK_DIR = getcwd()

CFG = load_config(f"{WRK_DIR}\\utils\\import_config.toml")

LOG = setup_logging(__name__,
                    CFG["general"]["log_level"],
                    log_dest=CFG["general"]["log_output"],
                    filename=f"{WRK_DIR}\\log_files\\ImportEvents_{datetime.now():%Y%m%d_%H%M%S}.log"
                    )

LOG.debug(f"Current Working Directory: {WRK_DIR}")

ansi = {
    "Black": "\u001b[30m",
    "Red": "\u001b[31m",
    "Green": "\u001b[32m",
    "Yellow": "\u001b[33m",
    "Blue": "\u001b[34m",
    "Magenta": "\u001b[35m",
    "Cyan": "\u001b[36m",
    "White": "\u001b[37m",
    "Bright Black": "\u001b[30;1m",
    "Bright Red": "\u001b[31;1m",
    "Bright Green": "\u001b[32;1m",
    "Bright Yellow": "\u001b[33;1m",
    "Bright Blue": "\u001b[34;1m",
    "Bright Magenta": "\u001b[35;1m",
    "Bright Cyan": "\u001b[36;1m",
    "Bright White": "\u001b[37;1m",
    "Bold": "\u001b[1m",
    "Underline": "\u001b[4m",
    "Reset": "\u001b[0m",
}

update_check_enabled = True

v = "v2.1.0"

t1 = perf_counter()


def install_lib(lib):
    """
    Installs the specified Python library using pip.

    Parameters:
        lib (str): The name of the library to install.
    """

    print(f"{ansi['Yellow']}\nInstalling {lib}...")
    # implement pip as a subprocess:
    check_call([executable, "-m", "pip", "install", lib])

    # process output with an API in the subprocess module:
    requests = check_output([executable, "-m", "pip", "freeze"])
    installed_packages = [r.decode().split("==")[0] for r in requests.split()]

    print(installed_packages)


# check if libraries are installed, if not, install it
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print(f"{ansi['Bright Red']}Openpyxl library is not installed.")
    install_lib("Openpyxl")
    from openpyxl import load_workbook
try:
    from requests import get
except ModuleNotFoundError:
    print(f"{ansi['Bright Red']}Requests library is not installed.")
    install_lib("requests")
    from requests import get
try:
    from tqdm import tqdm
except ModuleNotFoundError:
    print(f"{ansi['Bright Red']}tqdm library is not installed.")
    install_lib("tqdm")
    from tqdm import tqdm


def preamble():
    """
    Displays the script title, version, and optionally checks Python version and
    whether a newer version of the script is available on GitHub.
    """

    system("color")
    print(
        f"{ansi['Underline']}{ansi['Bright Magenta']}Events Layout Import Tool{ansi['Reset']}"
    )
    print(v)

    LOG.debug("Python Version: " + version[:7])

    if update_check_enabled:
        owner = "jkernal"
        repo = "EDC_Events_tool"
        print("Checking for updates...", end="", flush=True)
        try:
            url = f"https://api.github.com/repos/{owner}/{repo}/releases/latest"
            response = get(url, timeout=5)
            LOG.debug(f"Sent GET request to: {url}")
            LOG.debug(f"{response}")
            print("[DONE]")
            if v != response.json()["tag_name"]:
                print(
                    f"{ansi['Bright Yellow']}***There is a new release of this tool.***\nGo to: https://github.com/jkernal/EDC_Events_tool/releases"
                )
        except Exception as e:
            print("[FAILED]")
            LOG.warning(
                f"\u001b[33;1m***Warning: Could not connect to repository. Version check failed.***\n{e}"
            )


def manages_files():
    """
    Confirms the presence of template, input, and output directories and files.
    Copies the template file into the output directory for modification.

    Returns:
        list: File paths for the template, output copy, and input file.
    """

    temp_dir = f"{WRK_DIR}\\template"
    # confirming files
    try:
        LOG.debug(f"Template directory: {temp_dir}")
        temp_loc = f"{temp_dir}\\{listdir(temp_dir)[0]}"
        LOG.debug(f"Template location: {temp_loc}")
    except FileNotFoundError:
        LOG.error(f"Template directory not found. {temp_dir}")
        print(
            f"{ansi['Bold']}{ansi['Bright Red']}The template directory was not found.\n\nPlease add the template directory and restart."
        )
        done()
    except IndexError:
        LOG.error(f"Template file not found. {temp_dir}")
        print(
            f"{ansi['Bold']}{ansi['Bright Red']}The template file was not found.\n\nPlease add the template file to the template directory and restart."
        )
        done()
    # Copying template file to output directory
    try:
        output_path = f"{WRK_DIR}\\out_{listdir(temp_dir)[0]}"
        LOG.debug(f"Output file path: {output_path}")
        copy(temp_loc, output_path)
        LOG.info("Copied template file successfully.")
    except FileNotFoundError:
        print(
            f"{ansi['Bold']}{ansi['Bright Red']}The output directory was not found.\n\nPlease add the output directory and restart."
        )
        done()
    except Exception as e:
        LOG.error(f"Something went wrong copying the template file to {output_path}. {e}")
        print(
            f"{ansi['Bold']}{ansi['Bright Red']}Make sure to close the template file or make sure template file is not being used by another program."
        )
        done()

    toyo_loc = f"{WRK_DIR}\\loader\\bins\\toyo_comments.bin"

    toyo_path = Path(toyo_loc)
    LOG.debug(f"Toyopuc file location: {toyo_path}")

    if not toyo_path.exists():
        toyo_loc = None
        LOG.warning("Toyopuc binary file not found.")
    else:
        LOG.info("Toyopuc binary file found.")

    sw_loc = f"{WRK_DIR}\\loader\\bins\\sw_comments.bin"

    sw_path = Path(sw_loc)
    LOG.debug(f"ScreenWorks file location: {sw_path}")
    
    if not sw_path.exists():
        sw_loc = None
        LOG.warning("ScreenWorks binary file not found.")
    else:
        LOG.info("ScreenWorks binary file found.")

    locations = [temp_loc, output_path, toyo_loc, sw_loc]
    return locations


def perm_check(locs):
    """
    Checks read, write, and execute permissions for a list of files.

    Parameters:
        locs (list): List of file paths to check.
    """

    file_names = ["template", "output", "Toyopuc comment", "ScreenWorks"]
    access_type = ["read", "write", "execute"]
    for i in range(len(locs)):
        if locs[i] is None:
            continue
        permissions = [
            access(locs[i], R_OK),
            access(locs[i], W_OK),
            access(locs[i], X_OK),
        ]
        for j in range(len(permissions)):
            if not permissions[j]:
                print(
                    f"{ansi['Bold']}{ansi['Bright Red']}The script does not have {access_type[j]} access to the {file_names[i]} file. Make sure the file is closed and permissions are set."
                )
            else:
                continue
    return None


def get_address_array_from_template(sheet):
    """
    Extracts address values from column B of the Excel worksheet starting from row 3.

    Parameters:
        sheet (Worksheet): An openpyxl worksheet object.

    Returns:
        list: A list of [address, row number] pairs.
    """

    array = []
    for i in range(sheet.max_row):
        array.append([sheet.cell(i + 3, 2).value, i + 3])
    return array


def load_shm_as_dict(path, record_size=160):

    address_map = {}

    with open(path, "rb") as f:
        mm = mmap(f.fileno(), 0, access=ACCESS_READ)

        for i in range(0, len(mm), record_size):
            record = mm[i : i + record_size]
            addr = record[:64].decode("utf-8", errors="ignore").strip()
            comment = record[64:record_size].decode("utf-8", errors="ignore").strip()
            LOG.debug(f"From {path}: {addr} = {comment}")
            if addr:
                address_map[addr] = comment
        mm.close()
    return address_map


def done():
    """
    Resets terminal text formatting and exits the program after printing execution time.
    """

    print("\u001b[37m\u001b[0m")
    time_elapsed = round((perf_counter() - t1), 3)
    print(" ".join([f"{ansi['Reset']}Execution time: ", f"{time_elapsed}", "sec(s)"]))
    exit()


def remove_first_zero_if_long(s):
    if s is None:
        return
    if len(s) > 5:
        zero_index = s.find("0")
        if zero_index != -1 and zero_index < len(s) - 3:
            LOG.debug(f"Transformed address from {s} to {s[:zero_index] + s[zero_index + 1 :]}")
            return s[:zero_index] + s[zero_index + 1 :]
    return s


def is_valid_by_regex(s: str) -> bool:

    patterns = [r"FAULT\s?\d[A-Z]\d", r".*FAULT SPARE \d{1,2}$", r"(?:\b(?:TOTAL|OTHER)\s+)?FAULT(?:\s+\w+)?(?:\s+FAULT)?\s+\w+"]

    if not isinstance(s, str):
        LOG.debug(f"'{s}' is not a string. Rejected by regex.")
        return False  # or True if you want to skip None-like inputs
    
    for pattern in patterns:
        if fullmatch(pattern, s) and CFG["general"]["regex_match"] == "fullmatch":
            LOG.debug(f"'{s}' was rejected by regex pattern: {pattern}")
            return False  # it matched a disallowed pattern
        elif search(pattern, s) and CFG["general"]["regex_match"] == "search":
            LOG.debug(f"'{s}' was rejected by regex pattern: {pattern}")
            return False

    LOG.debug(f"'{s}' was valid by regex.")
    return True  # didn't match any patterns


def main():
    """
    Main execution flow:
    - Initializes script
    - Confirms and copies files
    - Checks file permissions
    - Extracts addresses from Excel and comments from CSV
    - Matches and writes comments to the Excel output
    - Saves file and reports results
    """

    # run preamble
    preamble()

    # run the address comment loader program
    loader_exe_path = Path(__file__).parent / "loader\\address_comment_loader.exe"
    LOG.debug(f"Loader exe location: {loader_exe_path}")
    #print(loader_exe_path)
    run(str(loader_exe_path))

    # find file locations
    file_locs = (
        manages_files()
    )  # file_locs[template, output, toyopuc bin, screenworks bin]

    # check permissions on files
    perm_check(file_locs)

    # open output workbook and worksheet
    wb = load_workbook(filename=file_locs[1])
    ws = wb["Import Cheat Sheet"]

    # wait for results from both threads
    lookup_addresses = get_address_array_from_template(ws)

    if file_locs[2] is None and file_locs[3] is None:
        LOG.error("No comment files were found.")
        print("No comment files were found. Please provide comment files and try again.")
        done()

    if file_locs[2] is not None:
        toyo_addr_comment_map = load_shm_as_dict(file_locs[2])
        toyo_exists = True
    else:
        toyo_exists = False

    if file_locs[3] is not None:
        sw_addr_comment_map = load_shm_as_dict(file_locs[3])
        sw_exists = True
    else:
        sw_exists = False

    toyo_match_count, sw_match_count, address_array_len = 0, 0, len(lookup_addresses)

    print(f"\n{ansi['Reset']}{ansi['Green']}Working on it...", flush=True, end="")

    # loop through the addresses and compare to the array with comments
    for i in tqdm(range(address_array_len)):
        
        if lookup_addresses[i][0] is None:
            continue
        elif len(lookup_addresses[i][0]) == 4:
            searched_address = "P1-" + lookup_addresses[i][0]
        else:
            searched_address = lookup_addresses[i][0]

        #print(f"From ScreenWorks: {searched_address} = {sw_addr_comment_map.get(searched_address)}")
        #print(f"From Toyopuc:  {searched_address} = {toyo_addr_comment_map.get(searched_address)}")
        
        if toyo_exists:
            toyo_result = toyo_addr_comment_map.get(searched_address)
            LOG.debug(f"Result from Toyopuc file: {searched_address} = {toyo_result}")
        else:
            toyo_result = None

        if sw_exists:
            sw_address = remove_first_zero_if_long(searched_address)
            sw_result = sw_addr_comment_map.get(sw_address)
            LOG.debug(f"Result from ScreenWorks file: {sw_address} = {sw_result}")
        else:
            sw_result = None

        if isinstance(toyo_result, str) and is_valid_by_regex(toyo_result):
            if toyo_result is not None:
                ws.cell(row=lookup_addresses[i][1], column=6).value = lookup_addresses[
                    i
                ][0]
                ws.cell(row=lookup_addresses[i][1], column=7).value = (
                    toyo_addr_comment_map.get(searched_address)
                )
                LOG.debug(f"{searched_address} comment is from Toyopuc.")
                toyo_match_count += 1
                continue
        
        elif isinstance(sw_result, str) and is_valid_by_regex(sw_result):
            if sw_result is not None:
                ws.cell(row=lookup_addresses[i][1], column=6).value = lookup_addresses[
                    i
                ][0]
                ws.cell(row=lookup_addresses[i][1], column=7).value = (
                    sw_addr_comment_map.get(searched_address)
                )
                LOG.debug(f"{searched_address} comment is from ScreenWorks.")
                sw_match_count += 1
                continue
        
        if toyo_result is not None:
            ws.cell(row=lookup_addresses[i][1], column=6).value = lookup_addresses[i][0]
            ws.cell(row=lookup_addresses[i][1], column=7).value = (
                toyo_addr_comment_map.get(searched_address)
            )
            LOG.debug(f"{searched_address} comment is from Toyopuc.")
            toyo_match_count += 1
        elif sw_result is not None:
            ws.cell(row=lookup_addresses[i][1], column=6).value = lookup_addresses[i][0]
            ws.cell(row=lookup_addresses[i][1], column=7).value = (
                sw_addr_comment_map.get(searched_address)
            )
            LOG.debug(f"{searched_address} comment is from ScreenWorks.")
            sw_match_count += 1

    # save changes to the output file
    wb.save(file_locs[1])
    LOG.info(f"Successfully wrote to {file_locs[1]}")

    # display stats and warning if needed
    print("\nDone.", flush=True)
    print(
        f"\n{ansi['Bright Blue']}Number of comments found in Toyopuc file:"
        f"{ansi['Yellow']}" + str(toyo_match_count)
    )
    print(
        f"{ansi['Bright Blue']}Number of comments found in the ScreenWorks file:"
        f"{ansi['Yellow']}" + str(sw_match_count)
    )
    if sw_match_count == 0 and toyo_match_count == 0:
        print(
            f"{ansi['Bright Yellow']}***No matches were found. "
            f"Make sure your input and template files are correct***"
        )

    # reset and exit()
    done()
    # end of main


# entry point
if __name__ == "__main__":
    main()
